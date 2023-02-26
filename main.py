#!/usr/bin/env python3

import csv
import sys
from collections import defaultdict
from pathlib import Path
from typing import Generator

import openpyxl

FilesPair = tuple[Path, Path]
DataPair = tuple[list[dict], list[dict]]

POSSIBLE_SUFFIXES: list[str] = [".xlsx", ".csv"]


def extract_files_from_args(argv: list[str]) -> tuple[Path, ...]:
    files: list[Path] = []
    for arg in argv:
        file: Path = Path(arg).expanduser().resolve()
        if not file.is_file():
            continue
        if not file.suffix.lower() in POSSIBLE_SUFFIXES:
            continue
        files.append(file)
    return tuple(files)


def extract_data_from_xlsx(file: Path) -> list[dict]:
    wb: openpyxl.Workbook = openpyxl.load_workbook(file)
    ws = wb.active
    rows: list[dict] = []
    columns: list[str] = []
    for i, row in enumerate(ws):
        if i == 0:
            columns = [str(cell.value) for cell in row]
            continue
        row_content: dict = {}
        for j, cell in enumerate(row):
            row_content[columns[j]] = cell.value
        rows.append(row_content)
    return rows


def extract_data_from_csv(file: Path) -> list[dict]:
    with open(file, "r", newline="") as f:
        return list(csv.DictReader(f))


def extract_data_from_files(files: FilesPair) -> DataPair:
    """Read data into list of dicts.

    .. code-block:: python

        [
            {A: A1, B: B1, ...}
            {A: A2, B: B2, ...}
        ]
    """

    def extract(file: Path) -> list[dict]:
        suffix: str = file.suffix.replace(".", "")
        extraction_function: callable = globals()["extract_data_from_" + suffix]
        spreadsheet_as_dict: list[dict] = extraction_function(file)
        return spreadsheet_as_dict

    file_a, file_b = files
    return extract(file_a), extract(file_b)


def check_table_structure_integrity(files: DataPair) -> None:
    file_a, file_b = files
    diff = file_a[0].keys() ^ file_b[0].keys()
    if diff:
        raise ValueError(f"Différence [{len(diff)} colonne(s)]: {', '.join(diff)}")


def _dict_of_lists_as_rows(dict_: dict) -> Generator[dict, None, None]:
    """Convert dict of lists to rows.

    .. code-block:: python

        _dict_of_lists_as_rows(
            {
                A: [A1, A2, ...],
                B: [B1, B2, ...],
            }
        )

        Generator([{A: A1, B: B1}, {A: A2, B: B2}, ...])
    """
    columns: tuple = tuple(dict_.keys())
    nb_rows: int = len(dict_[columns[0]])
    for i in range(nb_rows):
        row: dict = {}
        for column in columns:
            row[column] = dict_[column][i]
        yield row


def create_out_dir_if_not_exists() -> None:
    Path("out").mkdir(parents=True, exist_ok=True)


def _write_csv(file_name: str, data: list[dict]) -> None:
    with open(file_name, "w", newline="") as f:
        writer: csv.DictWriter = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        for row in data:
            writer.writerow(row)


def make_tab_1_file_a(files: DataPair) -> None:
    file_a, _ = files
    _write_csv("out/tab_1.csv", file_a)


def make_tab_2_file_b(files: DataPair) -> None:
    _, file_b = files
    _write_csv("out/tab_2.csv", file_b)


def _extract_difference(file_left: list[dict], file_right: list[dict]) -> list[dict]:
    difference: list[dict] = []
    for row in file_left:
        if row not in file_right:
            difference.append(row)
    return difference


def make_tab_3_in_file_a_but_not_in_b(files: DataPair) -> None:
    file_a, file_b = files
    difference: list[dict] = _extract_difference(file_a, file_b)
    _write_csv("out/tab_3.csv", difference)


def make_tab_4_in_file_b_but_not_in_a(files: DataPair) -> None:
    file_a, file_b = files
    difference: list[dict] = _extract_difference(file_b, file_a)
    _write_csv("out/tab_4.csv", difference)


def _extract_common_by_first_column(
    file_left: list[dict], file_right: list[dict]
) -> DataPair:
    common_left: list[dict] = []
    common_right: list[dict] = []
    first_column: str = list(file_left[0].keys())[0]
    for row_left, row_right in zip(file_left, file_right):
        if row_left[first_column] == row_right[first_column]:
            common_left.append(row_left)
            common_right.append(row_right)
    return common_left, common_right


def _build_column_headers(columns: list[str]) -> list[str]:
    headers: list[str] = []
    for column in columns:
        headers.append(column)
        headers.append(column)
        headers.append("Delta_" + column)
    return headers


def make_tab_5_analysis(files: DataPair) -> None:
    file_a, file_b = files
    common: DataPair = _extract_common_by_first_column(file_a, file_b)
    columns: list[str] = list(common[0][0].keys())
    headers: list[str] = _build_column_headers(columns)

    with open("out/tab_5.csv", "w", newline="") as f:
        writer: csv.writer = csv.writer(f)
        writer.writerow(headers)

        total_differences: defaultdict = defaultdict(int)
        for row_a, row_b in zip(*common):
            merged_row: list[str] = []
            for column in columns:
                row_a_value: str = row_a[column]
                row_b_value: str = row_b[column]
                merged_row.append(row_a_value)
                merged_row.append(row_b_value)
                is_different: int = int(row_a_value != row_b_value)
                merged_row.append(str(is_different))
                total_differences[column] += is_different
            writer.writerow(merged_row)

        totals_row: list[str] = []
        for column in columns:
            totals_row.append("")
            totals_row.append("")
            totals_row.append(str(total_differences[column]))
        writer.writerow(totals_row)


def csv_tab_files_to_excel_sheets() -> None:
    wb: openpyxl.Workbook = openpyxl.Workbook()
    for i, csv_file in enumerate(sorted(Path("out/").glob("*.csv"))):
        if i == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = f"Onglet {i+1}"

        with open(csv_file, "r", newline="") as f:
            reader: csv.reader = csv.reader(f)
            for row in reader:
                ws.append(row)

        csv_file.unlink()

    wb.save("out/out.xlsx")


def main(argv: list[str]) -> int:
    files: tuple[Path, ...] = extract_files_from_args(argv)
    if len(files) != 2:
        print(f"Erreur: 2 fichiers requis en entrée ({', '.join(POSSIBLE_SUFFIXES)})")
        return 2
    files: FilesPair

    files: DataPair = extract_data_from_files(files)
    try:
        check_table_structure_integrity(files)
    except ValueError as e:
        print(e)
        return 1

    create_out_dir_if_not_exists()

    make_tab_1_file_a(files)
    make_tab_2_file_b(files)
    make_tab_3_in_file_a_but_not_in_b(files)
    make_tab_4_in_file_b_but_not_in_a(files)
    make_tab_5_analysis(files)

    csv_tab_files_to_excel_sheets()

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
